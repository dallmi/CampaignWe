#!/usr/bin/env python3
"""
CampaignWe XLSX Report Generator

Generates a formatted, multi-tab Excel report from the processed Parquet files
to showcase campaign success metrics.

Usage:
    python generate_report.py                          # Default: read from output/, write to output/
    python generate_report.py --output report.xlsx     # Custom output path
    python generate_report.py --date-from 2026-01-01   # Filter by date range
    python generate_report.py --date-to 2026-03-31

Input:
    - output/events_anonymized.parquet
    - output/story_metadata.parquet

Output:
    - output/campaignwe_report.xlsx (9 tabs)
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Brand colors
# ---------------------------------------------------------------------------
CORP_RED = "E60000"
BORDEAUX_I = "BD000C"
LAKE_50 = "0C7EC6"
WHITE = "FFFFFF"
BLACK = "000000"
GRAY_I = "CCCABC"
GRAY_III = "8E8D83"
GRAY_IV = "7A7870"
GRAY_VI = "404040"
PASTEL_I = "ECEBE4"
ROW_ALT = "F8F7F2"
BRONZE_I = "B98E2C"
RAG_GREEN = "6F7A1A"
RAG_AMBER = "E4A911"

# ---------------------------------------------------------------------------
# Reusable styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=GRAY_VI, end_color=GRAY_VI, fill_type="solid")
SECTION_FONT = Font(bold=True, color=GRAY_VI, size=11)
SECTION_FILL = PatternFill(start_color=PASTEL_I, end_color=PASTEL_I, fill_type="solid")
ALT_FILL = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=PASTEL_I, end_color=PASTEL_I, fill_type="solid")
TOTAL_FONT = Font(bold=True, color=BLACK, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color=GRAY_I),
    right=Side(style="thin", color=GRAY_I),
    top=Side(style="thin", color=GRAY_I),
    bottom=Side(style="thin", color=GRAY_I),
)

NUM_FMT_INT = "#,##0"
NUM_FMT_PCT = "0.0%"
NUM_FMT_RATIO = "0.0"
NUM_FMT_DATE = "YYYY-MM-DD"


def log(msg=""):
    print(msg, flush=True)


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------
def write_header_row(ws, row, headers, col_start=1):
    """Write a styled header row."""
    for ci, h in enumerate(headers, start=col_start):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_data_rows(ws, start_row, rows, col_start=1, fmt_map=None):
    """Write data rows with alternating row color and number formats.

    fmt_map: dict mapping column-index (0-based within row) to an openpyxl number format string.
    """
    fmt_map = fmt_map or {}
    for ri, row_data in enumerate(rows):
        excel_row = start_row + ri
        fill = ALT_FILL if ri % 2 == 1 else None
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=excel_row, column=col_start + ci, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if ci in fmt_map:
                cell.number_format = fmt_map[ci]
            elif isinstance(val, float):
                cell.number_format = NUM_FMT_RATIO
            elif isinstance(val, int):
                cell.number_format = NUM_FMT_INT


def write_section_header(ws, row, text, col_span=2):
    """Write a section header spanning multiple columns."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    for ci in range(2, col_span + 1):
        c = ws.cell(row=row, column=ci)
        c.fill = SECTION_FILL
        c.border = THIN_BORDER


def write_kpi_row(ws, row, label, value, col_label=1, col_value=2, fmt=None):
    """Write a KPI label-value pair."""
    lbl = ws.cell(row=row, column=col_label, value=label)
    lbl.font = Font(bold=True, color=GRAY_VI)
    lbl.border = THIN_BORDER
    lbl.alignment = Alignment(indent=1)

    val = ws.cell(row=row, column=col_value, value=value)
    val.border = THIN_BORDER
    val.alignment = Alignment(horizontal="right")
    if fmt:
        val.number_format = fmt
    elif isinstance(value, float):
        val.number_format = NUM_FMT_RATIO
    elif isinstance(value, int):
        val.number_format = NUM_FMT_INT


def write_total_row(ws, row, data, col_start=1, fmt_map=None):
    """Write a bold total row."""
    fmt_map = fmt_map or {}
    for ci, val in enumerate(data):
        cell = ws.cell(row=row, column=col_start + ci, value=val)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if ci in fmt_map:
            cell.number_format = fmt_map[ci]
        elif isinstance(val, float):
            cell.number_format = NUM_FMT_RATIO
        elif isinstance(val, int):
            cell.number_format = NUM_FMT_INT


def write_formula(ws, row, col, formula, fmt=None, fill=None, bold=False):
    """Write an Excel formula into a cell with styling."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if bold:
        cell.font = TOTAL_FONT


def auto_fit_columns(ws, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def finalize_sheet(ws, freeze_row=2):
    """Freeze panes and finalize a sheet."""
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_data(events_path, metadata_path, date_from=None, date_to=None):
    """Load parquet files into an in-memory DuckDB and return the connection."""
    con = duckdb.connect(":memory:")

    if not events_path.exists():
        log(f"ERROR: Events file not found: {events_path}")
        sys.exit(1)

    con.execute(f"CREATE TABLE events AS SELECT * FROM read_parquet('{events_path}')")
    log(f"  Loaded {con.execute('SELECT COUNT(*) FROM events').fetchone()[0]:,} events")

    # Check available columns
    cols = [r[0] for r in con.execute("DESCRIBE events").fetchall()]

    if metadata_path.exists():
        con.execute(f"CREATE TABLE story_meta AS SELECT * FROM read_parquet('{metadata_path}')")
        log(f"  Loaded {con.execute('SELECT COUNT(*) FROM story_meta').fetchone()[0]:,} stories")
    else:
        log(f"  WARNING: Story metadata not found: {metadata_path}")
        con.execute("CREATE TABLE story_meta (story_id VARCHAR, story_title VARCHAR, status VARCHAR)")

    # Apply date filters
    if date_from:
        con.execute(f"DELETE FROM events WHERE session_date < '{date_from}'")
        log(f"  Filtered: session_date >= {date_from}")
    if date_to:
        con.execute(f"DELETE FROM events WHERE session_date > '{date_to}'")
        log(f"  Filtered: session_date <= {date_to}")

    remaining = con.execute("SELECT COUNT(*) FROM events").fetchone()[0]
    log(f"  {remaining:,} events after date filters")

    return con, cols


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------
def build_executive_summary(wb, con, cols):
    """Tab 1: Executive Summary — KPIs at a glance."""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_properties.tabColor = BRONZE_I

    kpis = con.execute("""
        SELECT
            MIN(session_date) as first_date,
            MAX(session_date) as last_date,
            DATEDIFF('day', MIN(session_date), MAX(session_date)) + 1
              - 2 * (DATEDIFF('week', MIN(session_date), MAX(session_date)))
              - CASE WHEN DAYOFWEEK(MIN(session_date)) = 1 THEN 1 ELSE 0 END
              - CASE WHEN DAYOFWEEK(MAX(session_date)) = 7 THEN 1 ELSE 0 END
              as duration_days,
            COUNT(CASE WHEN action_type != 'Cancel' THEN 1 END) as total_interactions,
            COUNT(DISTINCT person_hash) as unique_visitors,
            COUNT(DISTINCT session_key) as unique_sessions,
            COUNT(DISTINCT CASE WHEN story_id IS NOT NULL THEN story_id END) as total_stories,
            COUNT(CASE WHEN action_type = 'Read' THEN 1 END) as reads,
            COUNT(CASE WHEN action_type = 'Like' THEN 1 END) as likes,
            COUNT(CASE WHEN action_type = 'Submit' THEN 1 END) as submits,
            COUNT(CASE WHEN action_type = 'Open Form' THEN 1 END) as open_forms,
            COUNT(CASE WHEN action_type = 'Cancel' THEN 1 END) as cancels,
            COUNT(CASE WHEN action_type = 'Delete' THEN 1 END) as deletes,
            COUNT(CASE WHEN action_type = 'Send Invite' THEN 1 END) as invites_sent,
            COUNT(CASE WHEN action_type = 'Open Invite' THEN 1 END) as invites_opened
        FROM events
    """).fetchone()

    first_date, last_date, duration, total, uv, sessions, stories = kpis[:7]
    reads, likes, submits, open_forms, cancels, deletes, invites_sent, invites_opened = kpis[7:]

    # Story counts from metadata
    active_stories = con.execute("""
        SELECT COUNT(DISTINCT story_id) FROM story_meta WHERE status = 'active'
    """).fetchone()[0]
    deleted_stories = con.execute("""
        SELECT COUNT(DISTINCT story_id) FROM story_meta WHERE status = 'deleted'
    """).fetchone()[0]
    pending_stories = con.execute("""
        SELECT COUNT(DISTINCT story_id) FROM story_meta WHERE status = 'pending'
    """).fetchone()[0]

    # Aggregate click categories
    engagement_clicks = reads + likes
    invite_clicks = invites_sent + invites_opened
    submission_clicks = open_forms + submits + deletes

    # --- Layout --- (track row numbers for formula references)
    r = 1
    write_section_header(ws, r, "CAMPAIGN OVERVIEW", 2); r += 1
    write_kpi_row(ws, r, "Report Period", f"{first_date} to {last_date}"); r += 1
    write_kpi_row(ws, r, "Duration (business days)", duration, fmt=NUM_FMT_INT); row_dur = r; r += 1
    r += 1

    write_section_header(ws, r, "REACH", 2); r += 1
    write_kpi_row(ws, r, "Total Clicks", total, fmt=NUM_FMT_INT); row_clicks = r; r += 1
    # Click breakdown — % share baked into label via formula, smaller italic font
    SUB_FONT = Font(italic=True, color=GRAY_IV, size=10)
    breakdown = [
        (r, "Engagement (Read + Like)", engagement_clicks),
        (r + 1, "Invite (Open + Send)", invite_clicks),
        (r + 2, "Submission (Form + Submit + Delete)", submission_clicks),
    ]
    for br_row, label, count in breakdown:
        # Column A: formula that concatenates the % with the label
        cell_a = ws.cell(row=br_row, column=1)
        cell_a.value = f'=TEXT(IF(B${row_clicks}=0,0,B{br_row}/B${row_clicks}),"0%") & "  {label}"'
        cell_a.font = SUB_FONT
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(indent=4)
        # Column B: raw count
        cell_b = ws.cell(row=br_row, column=2, value=count)
        cell_b.font = SUB_FONT
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal="right")
        cell_b.number_format = NUM_FMT_INT
    r += 3
    write_kpi_row(ws, r, "Unique Visitors", uv, fmt=NUM_FMT_INT); row_uv = r; r += 1
    write_kpi_row(ws, r, "Unique Sessions", sessions, fmt=NUM_FMT_INT); r += 1
    # Formula: Total Clicks / Unique Visitors
    ws.cell(row=r, column=1, value="Avg. Clicks / Visitor").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_uv}=0,0,B{row_clicks}/B{row_uv})", fmt=NUM_FMT_RATIO); r += 1
    # Formula: Unique Visitors / Duration
    ws.cell(row=r, column=1, value="Avg. Daily Visitors").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_dur}=0,0,B{row_uv}/B{row_dur})", fmt=NUM_FMT_RATIO); r += 1
    r += 1

    write_section_header(ws, r, "INTERACTION", 2); r += 1
    write_kpi_row(ws, r, "Story Reads", reads, fmt=NUM_FMT_INT); row_reads = r; r += 1
    write_kpi_row(ws, r, "Story Likes", likes, fmt=NUM_FMT_INT); row_likes = r; r += 1
    # Formula: Likes / Reads
    ws.cell(row=r, column=1, value="Like Rate (Likes / Reads)").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_reads}=0,0,B{row_likes}/B{row_reads})", fmt=NUM_FMT_PCT); r += 1
    write_kpi_row(ws, r, "Invites Opened", invites_opened, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Invites Sent", invites_sent, fmt=NUM_FMT_INT); r += 1
    r += 1

    write_section_header(ws, r, "CONTENT", 2); r += 1
    write_kpi_row(ws, r, "Total Stories (all time)", stories, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Active Stories", active_stories, fmt=NUM_FMT_INT); row_active = r; r += 1
    write_kpi_row(ws, r, "Deleted Stories", deleted_stories, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Pending / Unapproved Stories", pending_stories, fmt=NUM_FMT_INT); r += 1
    # Formula: Reads / Active Stories
    ws.cell(row=r, column=1, value="Avg. Reads / Active Story").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_active}=0,0,B{row_reads}/B{row_active})", fmt=NUM_FMT_RATIO); r += 1
    # Formula: Likes / Active Stories
    ws.cell(row=r, column=1, value="Avg. Likes / Active Story").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_active}=0,0,B{row_likes}/B{row_active})", fmt=NUM_FMT_RATIO); r += 1
    r += 1

    write_section_header(ws, r, "SUBMISSION FUNNEL", 2); r += 1
    write_kpi_row(ws, r, "Opened Form", open_forms, fmt=NUM_FMT_INT); row_openform = r; r += 1
    write_kpi_row(ws, r, "Submitted", submits, fmt=NUM_FMT_INT); row_submit = r; r += 1
    # Formula: Submits / Open Forms
    ws.cell(row=r, column=1, value="Submission Rate (Submit / Open Form)").font = Font(bold=True, color=GRAY_VI)
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=1).alignment = Alignment(indent=1)
    write_formula(ws, r, 2, f"=IF(B{row_openform}=0,0,B{row_submit}/B{row_openform})", fmt=NUM_FMT_PCT); r += 1
    write_kpi_row(ws, r, "Delete Confirmations (Clicks)", deletes, fmt=NUM_FMT_INT); r += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = ws.cell(row=2, column=1)
    log("  Tab 1: Executive Summary")


def build_weekly_trend(wb, con):
    """Tab 2: Weekly Trend — growth and momentum."""
    ws = wb.create_sheet("Weekly Trend")
    ws.sheet_properties.tabColor = GRAY_IV

    rows = con.execute("""
        WITH weekly AS (
            SELECT
                YEARWEEK(session_date) as yw,
                MIN(session_date) as week_start,
                COUNT(CASE WHEN action_type != 'Cancel' THEN 1 END) as clicks,
                COUNT(DISTINCT person_hash) as uv,
                COUNT(CASE WHEN action_type = 'Read' THEN 1 END) as reads,
                COUNT(CASE WHEN action_type = 'Like' THEN 1 END) as likes,
                COUNT(CASE WHEN action_type = 'Submit' THEN 1 END) as submits,
                COUNT(CASE WHEN action_type = 'Open Form' THEN 1 END) as open_forms,
                COUNT(CASE WHEN action_type = 'Send Invite' THEN 1 END) as invites_sent,
                COUNT(CASE WHEN action_type = 'Open Invite' THEN 1 END) as invites_opened,
                COUNT(CASE WHEN action_type = 'Delete' THEN 1 END) as deletes,
                COUNT(DISTINCT session_date) as active_days
            FROM events
            GROUP BY YEARWEEK(session_date)
        ),
        first_seen AS (
            SELECT person_hash, MIN(YEARWEEK(session_date)) as first_week
            FROM events
            GROUP BY person_hash
        ),
        new_per_week AS (
            SELECT first_week as yw, COUNT(*) as new_visitors
            FROM first_seen
            GROUP BY first_week
        )
        SELECT
            w.yw,
            w.week_start,
            w.clicks,
            w.uv,
            COALESCE(n.new_visitors, 0) as new_visitors,
            w.uv - COALESCE(n.new_visitors, 0) as returning_visitors,
            w.reads,
            w.likes,
            w.submits,
            w.open_forms,
            w.invites_opened,
            w.invites_sent,
            w.deletes
        FROM weekly w
        LEFT JOIN new_per_week n ON w.yw = n.yw
        ORDER BY w.yw
    """).fetchall()

    # Col: A=Week B=Start C=Clicks D=UV E=New F=Return G=Reads H=Likes
    #      I=Submit J=OpenForm K=InvOpen L=InvSent M=Deletes N=LikeRate
    headers = [
        "Week", "Week Start", "Clicks", "Unique Visitors",
        "New Visitors", "Returning Visitors", "Reads", "Likes",
        "Submissions", "Open Form", "Invites Opened",
        "Invites Sent", "Deletes", "Like Rate"
    ]
    fmt = {
        0: "0", 1: NUM_FMT_DATE, 2: NUM_FMT_INT, 3: NUM_FMT_INT,
        4: NUM_FMT_INT, 5: NUM_FMT_INT, 6: NUM_FMT_INT, 7: NUM_FMT_INT,
        8: NUM_FMT_INT, 9: NUM_FMT_INT, 10: NUM_FMT_INT, 11: NUM_FMT_INT,
        12: NUM_FMT_INT
    }

    write_header_row(ws, 1, headers)
    write_data_rows(ws, 2, rows, fmt_map=fmt)

    # Like Rate as formula: =IF(G{r}=0,0,H{r}/G{r})  (Likes/Reads)
    for ri in range(len(rows)):
        r = ri + 2
        write_formula(ws, r, 14, f"=IF(G{r}=0,0,H{r}/G{r})", fmt=NUM_FMT_PCT,
                      fill=ALT_FILL if ri % 2 == 1 else None)

    # Data bars on Clicks and UV columns
    if rows:
        last_row = len(rows) + 1
        for col_letter in ["C", "D"]:
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                DataBarRule(start_type="num", start_value=0, end_type="max",
                            color=GRAY_III, showValue=True)
            )

    finalize_sheet(ws)
    log("  Tab 2: Weekly Trend")


def build_story_performance(wb, con):
    """Tab 3: Story Performance — which stories resonate most."""
    ws = wb.create_sheet("Story Performance")
    ws.sheet_properties.tabColor = GRAY_IV

    rows = con.execute("""
        WITH story_events AS (
            SELECT
                e.story_id,
                COUNT(CASE WHEN e.action_type = 'Read' THEN 1 END) as total_reads,
                COUNT(DISTINCT CASE WHEN e.action_type = 'Read' THEN e.person_hash END) as unique_readers,
                COUNT(CASE WHEN e.action_type = 'Like' THEN 1 END) as likes,
                MEDIAN(CASE WHEN e.action_type = 'Read' THEN e.read_duration_sec END) as median_read_duration
            FROM events e
            WHERE e.story_id IS NOT NULL
            GROUP BY e.story_id
        )
        SELECT
            se.story_id,
            COALESCE(m.story_title, '(unknown)') as title,
            COALESCE(m.keys, '') as keys,
            COALESCE(m.author_division, '') as author_division,
            COALESCE(m.author_region, '') as author_region,
            COALESCE(m.status, 'unknown') as status,
            m.created,
            se.total_reads,
            se.unique_readers,
            se.likes,
            CASE WHEN m.created IS NOT NULL THEN
                DATEDIFF('day', m.created::DATE, COALESCE(m.deleted_date, CURRENT_DATE)) + 1
                - 2 * DATEDIFF('week', m.created::DATE, COALESCE(m.deleted_date, CURRENT_DATE))
                - CASE WHEN DAYOFWEEK(m.created::DATE) = 1 THEN 1 ELSE 0 END
                - CASE WHEN DAYOFWEEK(COALESCE(m.deleted_date, CURRENT_DATE)) = 7 THEN 1 ELSE 0 END
            ELSE NULL END as lifespan_days,
            ROUND(se.median_read_duration, 1) as median_read_duration
        FROM story_events se
        LEFT JOIN story_meta m ON se.story_id = m.story_id
        ORDER BY se.total_reads DESC
    """).fetchall()

    # Col: A=ID B=Title C=Keys D=AuthDiv E=AuthRegion F=Status G=Created H=Reads
    #      I=UniqueReaders J=Likes K=LikeRate L=Lifespan M=Reads/Day N=MedianReadDur
    headers = [
        "Story ID", "Title", "Keys", "Author Division", "Author Region",
        "Status", "Created",
        "Total Reads", "Unique Readers", "Likes", "Like Rate",
        "Lifespan (days)", "Reads/Day", "Median Read Duration (s)"
    ]
    fmt = {
        6: NUM_FMT_DATE, 7: NUM_FMT_INT, 8: NUM_FMT_INT, 9: NUM_FMT_INT,
        11: NUM_FMT_INT
    }

    write_header_row(ws, 1, headers)
    # Write data without the formula columns (Like Rate=col K, Reads/Day=col M)
    for ri, row_data in enumerate(rows):
        r = ri + 2
        # Write cols A-J (indices 0-9)
        data_cells = list(row_data[:10])  # ID through Likes
        fill = ALT_FILL if ri % 2 == 1 else None
        for ci, val in enumerate(data_cells):
            cell = ws.cell(row=r, column=ci + 1, value=val)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if ci in fmt:
                cell.number_format = fmt[ci]
            elif isinstance(val, int):
                cell.number_format = NUM_FMT_INT
        # Like Rate formula: =IF(I{r}=0,0,J{r}/I{r})  (Likes/Unique Readers)
        write_formula(ws, r, 11, f"=IF(I{r}=0,0,J{r}/I{r})", fmt=NUM_FMT_PCT, fill=fill)
        # Lifespan (col L, index 10 in SQL result)
        cell = ws.cell(row=r, column=12, value=row_data[10])
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_INT
        if fill:
            cell.fill = fill
        # Reads/Day formula: =IF(L{r}=0,"",H{r}/L{r})  (Total Reads/Lifespan)
        write_formula(ws, r, 13, f'=IF(L{r}=0,"",H{r}/L{r})', fmt=NUM_FMT_RATIO, fill=fill)
        # Median Read Duration (col N, index 11 in SQL result)
        cell = ws.cell(row=r, column=14, value=row_data[11])
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_RATIO
        if fill:
            cell.fill = fill

    finalize_sheet(ws)
    log("  Tab 3: Story Performance")


def build_key_performance(wb, con):
    """Tab 4: Key Performance — which story keys drive the most engagement."""
    ws = wb.create_sheet("3Keys Performance")
    ws.sheet_properties.tabColor = GRAY_IV

    # Check if key columns exist in story_meta
    meta_cols = [r[0] for r in con.execute("DESCRIBE story_meta").fetchall()]
    has_split_keys = all(f"story_key{i}" in meta_cols for i in range(1, 4))
    has_keys = "keys" in meta_cols

    if not has_keys and not has_split_keys:
        ws.cell(row=1, column=1, value="No key data available in story metadata")
        log("  Tab 4: Key Performance (skipped — no key data)")
        return

    # Unpivot keys: each story can have up to 3 keys, we need one row per key
    if has_split_keys:
        key_union = """
            SELECT story_id, TRIM(story_key1) as key FROM story_meta WHERE status != 'pending' AND TRIM(COALESCE(story_key1, '')) != ''
            UNION ALL
            SELECT story_id, TRIM(story_key2) FROM story_meta WHERE status != 'pending' AND TRIM(COALESCE(story_key2, '')) != ''
            UNION ALL
            SELECT story_id, TRIM(story_key3) FROM story_meta WHERE status != 'pending' AND TRIM(COALESCE(story_key3, '')) != ''
        """
    else:
        key_union = """
            SELECT story_id, TRIM(UNNEST(string_split(keys, ','))) as key
            FROM story_meta
            WHERE status != 'pending' AND keys IS NOT NULL AND TRIM(keys) != ''
        """

    rows = con.execute(f"""
        WITH story_keys AS ({key_union}),
        key_engagement AS (
            SELECT
                sk.key,
                COUNT(DISTINCT sk.story_id) as stories,
                COUNT(CASE WHEN e.action_type = 'Read' THEN 1 END) as reads,
                COUNT(DISTINCT CASE WHEN e.action_type = 'Read' THEN e.person_hash END) as unique_readers,
                COUNT(CASE WHEN e.action_type = 'Like' THEN 1 END) as likes,
                COUNT(CASE WHEN e.action_type IN ('Read', 'Like') THEN 1 END) as interactions,
                COUNT(DISTINCT CASE WHEN e.action_type IN ('Read', 'Like') THEN e.person_hash END) as active_visitors
            FROM story_keys sk
            LEFT JOIN events e ON sk.story_id = e.story_id
                AND e.action_type IN ('Read', 'Like')
            GROUP BY sk.key
        )
        SELECT
            key,
            stories,
            interactions,
            active_visitors,
            reads,
            likes
        FROM key_engagement
        WHERE key IS NOT NULL AND TRIM(key) != ''
        ORDER BY interactions DESC
    """).fetchall()

    # Col: A=Key B=Stories C=Interactions D=ActiveVisitors E=Reads F=Likes
    #      G=LikeRate H=AvgReads/Story I=AvgLikes/Story
    headers = [
        "Key", "Stories", "Interactions", "Active Visitors", "Reads", "Likes",
        "Like Rate", "Avg. Reads / Story", "Avg. Likes / Story"
    ]
    fmt_data = {
        1: NUM_FMT_INT, 2: NUM_FMT_INT, 3: NUM_FMT_INT, 4: NUM_FMT_INT, 5: NUM_FMT_INT
    }

    write_header_row(ws, 1, headers)
    write_data_rows(ws, 2, rows, fmt_map=fmt_data)

    # Formula columns
    for ri in range(len(rows)):
        r = ri + 2
        fill = ALT_FILL if ri % 2 == 1 else None
        write_formula(ws, r, 7, f"=IF(E{r}=0,0,F{r}/E{r})", fmt=NUM_FMT_PCT, fill=fill)
        write_formula(ws, r, 8, f"=IF(B{r}=0,0,E{r}/B{r})", fmt=NUM_FMT_RATIO, fill=fill)
        write_formula(ws, r, 9, f"=IF(B{r}=0,0,F{r}/B{r})", fmt=NUM_FMT_RATIO, fill=fill)

    finalize_sheet(ws)
    log("  Tab 4: Key Performance")


def build_division_engagement(wb, con, cols):
    """Tab 4: Division Interactions."""
    ws = wb.create_sheet("Division Interactions")
    ws.sheet_properties.tabColor = GRAY_IV

    if "visitor_division" not in cols:
        ws.cell(row=1, column=1, value="No organizational data available (visitor_division missing)")
        log("  Tab 4: Division Interactions (skipped — no org data)")
        return

    rows = con.execute("""
        SELECT
            COALESCE(visitor_division, '(Unknown)') as division,
            COUNT(DISTINCT person_hash) as active_visitors,
            COUNT(*) as interactions,
            COUNT(CASE WHEN action_type = 'Read' THEN 1 END) as reads,
            COUNT(CASE WHEN action_type = 'Like' THEN 1 END) as likes
        FROM events
        WHERE action_type IN ('Read', 'Like')
        GROUP BY COALESCE(visitor_division, '(Unknown)')
        ORDER BY active_visitors DESC
    """).fetchall()

    # Col: A=Division B=ActiveVisitors C=Interactions D=Reads E=Likes
    #      F=Int/Visitor G=LikeRate H=%ofTotal
    headers = [
        "Division", "Active Visitors", "Interactions", "Reads", "Likes",
        "Interactions / Visitor", "Like Rate", "% of Total"
    ]
    fmt_data = {1: NUM_FMT_INT, 2: NUM_FMT_INT, 3: NUM_FMT_INT, 4: NUM_FMT_INT}

    write_header_row(ws, 1, headers)
    write_data_rows(ws, 2, rows, fmt_map=fmt_data)

    total_row = len(rows) + 2

    # Formula columns for each data row
    for ri in range(len(rows)):
        r = ri + 2
        fill = ALT_FILL if ri % 2 == 1 else None
        write_formula(ws, r, 6, f"=IF(B{r}=0,0,C{r}/B{r})", fmt=NUM_FMT_RATIO, fill=fill)
        write_formula(ws, r, 7, f"=IF(D{r}=0,0,E{r}/D{r})", fmt=NUM_FMT_PCT, fill=fill)
        write_formula(ws, r, 8, f"=IF(B${total_row}=0,0,B{r}/B${total_row})", fmt=NUM_FMT_PCT, fill=fill)

    # Total row with SUM formulas
    last_data = total_row - 1
    total_data = ["TOTAL"]
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for ci in range(2, 6):  # B-E: SUM
        col_l = get_column_letter(ci)
        write_formula(ws, total_row, ci, f"=SUM({col_l}2:{col_l}{last_data})",
                      fmt=NUM_FMT_INT, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 6, f"=IF(B{total_row}=0,0,C{total_row}/B{total_row})",
                  fmt=NUM_FMT_RATIO, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 7, f"=IF(D{total_row}=0,0,E{total_row}/D{total_row})",
                  fmt=NUM_FMT_PCT, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 8, "=1", fmt=NUM_FMT_PCT, fill=TOTAL_FILL, bold=True)

    finalize_sheet(ws)
    log("  Tab 4: Division Interactions")


def build_region_engagement(wb, con, cols):
    """Tab 5: Region Interactions."""
    ws = wb.create_sheet("Region Interactions")
    ws.sheet_properties.tabColor = GRAY_IV

    if "visitor_region" not in cols:
        ws.cell(row=1, column=1, value="No regional data available (visitor_region missing)")
        log("  Tab 5: Region Interactions (skipped — no region data)")
        return

    rows = con.execute("""
        SELECT
            COALESCE(visitor_region, '(Unknown)') as region,
            COUNT(DISTINCT person_hash) as active_visitors,
            COUNT(*) as interactions,
            COUNT(CASE WHEN action_type = 'Read' THEN 1 END) as reads,
            COUNT(CASE WHEN action_type = 'Like' THEN 1 END) as likes
        FROM events
        WHERE action_type IN ('Read', 'Like')
        GROUP BY COALESCE(visitor_region, '(Unknown)')
        ORDER BY active_visitors DESC
    """).fetchall()

    # Col: A=Region B=ActiveVisitors C=Interactions D=Reads E=Likes
    #      F=Int/Visitor G=LikeRate H=%ofTotal
    headers = [
        "Region", "Active Visitors", "Interactions", "Reads", "Likes",
        "Interactions / Visitor", "Like Rate", "% of Total"
    ]
    fmt_data = {1: NUM_FMT_INT, 2: NUM_FMT_INT, 3: NUM_FMT_INT, 4: NUM_FMT_INT}

    write_header_row(ws, 1, headers)
    write_data_rows(ws, 2, rows, fmt_map=fmt_data)

    total_row = len(rows) + 2

    for ri in range(len(rows)):
        r = ri + 2
        fill = ALT_FILL if ri % 2 == 1 else None
        write_formula(ws, r, 6, f"=IF(B{r}=0,0,C{r}/B{r})", fmt=NUM_FMT_RATIO, fill=fill)
        write_formula(ws, r, 7, f"=IF(D{r}=0,0,E{r}/D{r})", fmt=NUM_FMT_PCT, fill=fill)
        write_formula(ws, r, 8, f"=IF(B${total_row}=0,0,B{r}/B${total_row})", fmt=NUM_FMT_PCT, fill=fill)

    # Total row with SUM formulas
    last_data = total_row - 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for ci in range(2, 6):
        col_l = get_column_letter(ci)
        write_formula(ws, total_row, ci, f"=SUM({col_l}2:{col_l}{last_data})",
                      fmt=NUM_FMT_INT, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 6, f"=IF(B{total_row}=0,0,C{total_row}/B{total_row})",
                  fmt=NUM_FMT_RATIO, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 7, f"=IF(D{total_row}=0,0,E{total_row}/D{total_row})",
                  fmt=NUM_FMT_PCT, fill=TOTAL_FILL, bold=True)
    write_formula(ws, total_row, 8, "=1", fmt=NUM_FMT_PCT, fill=TOTAL_FILL, bold=True)

    finalize_sheet(ws)
    log("  Tab 5: Region Interactions")


def build_hourly_weekday(wb, con):
    """Tab 6: Hourly & Weekday patterns with heatmap."""
    ws = wb.create_sheet("Hourly & Weekday")
    ws.sheet_properties.tabColor = GRAY_IV

    # All three parts filtered to engagement = Read + Like
    ENG_FILTER = "WHERE action_type IN ('Read', 'Like')"

    # Part A: Weekday summary
    r = 1
    write_section_header(ws, r, "WEEKDAY SUMMARY (Interaction: Reads + Likes)", 6); r += 1
    weekday_headers = ["Weekday", "Interactions", "Active Visitors", "Reads", "Likes", "Avg. per Day"]
    write_header_row(ws, r, weekday_headers); r += 1

    weekday_rows = con.execute(f"""
        SELECT
            event_weekday,
            COUNT(*) as engagements,
            COUNT(DISTINCT person_hash) as uv,
            COUNT(CASE WHEN action_type = 'Read' THEN 1 END) as reads,
            COUNT(CASE WHEN action_type = 'Like' THEN 1 END) as likes,
            ROUND(COUNT(*) * 1.0 / NULLIF(COUNT(DISTINCT session_date), 0), 1) as avg_per_day
        FROM events
        {ENG_FILTER}
        GROUP BY event_weekday, event_weekday_num
        ORDER BY event_weekday_num
    """).fetchall()

    fmt_wd = {1: NUM_FMT_INT, 2: NUM_FMT_INT, 3: NUM_FMT_INT, 4: NUM_FMT_INT, 5: NUM_FMT_RATIO}
    write_data_rows(ws, r, weekday_rows, fmt_map=fmt_wd)
    r += len(weekday_rows) + 2

    # Part B: Hourly summary
    write_section_header(ws, r, "HOURLY SUMMARY (Interaction, CET)", 4); r += 1
    hourly_headers = ["Hour (CET)", "Interactions", "Active Visitors"]
    write_header_row(ws, r, hourly_headers); r += 1

    hourly_rows = con.execute(f"""
        SELECT
            event_hour,
            COUNT(*) as engagements,
            COUNT(DISTINCT person_hash) as uv
        FROM events
        {ENG_FILTER}
        GROUP BY event_hour
        ORDER BY event_hour
    """).fetchall()

    fmt_h = {0: "00", 1: NUM_FMT_INT, 2: NUM_FMT_INT}
    write_data_rows(ws, r, hourly_rows, fmt_map=fmt_h)
    r += len(hourly_rows) + 2

    # Part C: Heatmap matrix (weekday x hour)
    write_section_header(ws, r, "INTERACTION HEATMAP (Reads + Likes by Weekday x Hour CET)", 26); r += 1

    heatmap_headers = [""] + [f"{h:02d}" for h in range(24)]
    write_header_row(ws, r, heatmap_headers); r += 1

    heatmap_data = con.execute(f"""
        SELECT event_weekday, event_weekday_num, event_hour, COUNT(*) as cnt
        FROM events
        {ENG_FILTER}
        GROUP BY event_weekday, event_weekday_num, event_hour
        ORDER BY event_weekday_num, event_hour
    """).fetchall()

    # Build matrix
    weekdays_ordered = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    matrix = {wd: [0] * 24 for wd in weekdays_ordered}
    for wd, _, hour, cnt in heatmap_data:
        if wd in matrix:
            matrix[wd][int(hour)] = cnt

    heatmap_start_row = r
    for wd in weekdays_ordered:
        row_data = [wd] + matrix[wd]
        for ci, val in enumerate(row_data):
            cell = ws.cell(row=r, column=ci + 1, value=val)
            cell.border = THIN_BORDER
            if ci == 0:
                cell.font = Font(bold=True)
            else:
                cell.number_format = NUM_FMT_INT
                cell.alignment = Alignment(horizontal="center")
        r += 1

    # Color scale on heatmap cells
    heatmap_end_row = r - 1
    heatmap_range = f"B{heatmap_start_row}:Y{heatmap_end_row}"
    ws.conditional_formatting.add(
        heatmap_range,
        ColorScaleRule(
            start_type="min", start_color=WHITE,
            mid_type="percentile", mid_value=50, mid_color=RAG_AMBER,
            end_type="max", end_color=CORP_RED
        )
    )

    finalize_sheet(ws, freeze_row=1)
    log("  Tab 6: Hourly & Weekday")


def _write_funnel(ws, r, title, steps):
    """Write a funnel section with formula-based percentages. Returns next free row."""
    write_section_header(ws, r, title, 4); r += 1
    headers = ["Step", "Unique Visitors", "% of Previous Step", "% of First Step"]
    write_header_row(ws, r, headers); r += 1

    first_data_row = r
    for i, (label, count) in enumerate(steps):
        fill = ALT_FILL if i % 2 == 1 else None
        cell_a = ws.cell(row=r, column=1, value=label)
        cell_a.border = THIN_BORDER
        if fill:
            cell_a.fill = fill
        cell_b = ws.cell(row=r, column=2, value=count)
        cell_b.border = THIN_BORDER
        cell_b.number_format = NUM_FMT_INT
        if fill:
            cell_b.fill = fill
        # Conversion Rate: % of previous step
        if i == 0:
            write_formula(ws, r, 3, "=1", fmt=NUM_FMT_PCT, fill=fill)
        else:
            write_formula(ws, r, 3, f"=IF(B{r-1}=0,0,B{r}/B{r-1})",
                          fmt=NUM_FMT_PCT, fill=fill)
        # % of first step
        write_formula(ws, r, 4, f"=IF(B${first_data_row}=0,0,B{r}/B${first_data_row})",
                      fmt=NUM_FMT_PCT, fill=fill)
        r += 1

    return r


def build_conversion_funnel(wb, con):
    """Tab 7: Conversion Funnels — three separate user journeys."""
    ws = wb.create_sheet("Conversion Funnels")
    ws.sheet_properties.tabColor = GRAY_IV

    funnel = con.execute("""
        SELECT
            COUNT(DISTINCT person_hash) as total_visitors,
            COUNT(DISTINCT CASE WHEN action_type = 'Read' THEN person_hash END) as readers,
            COUNT(DISTINCT CASE WHEN action_type = 'Like' THEN person_hash END) as likers,
            COUNT(DISTINCT CASE WHEN action_type = 'Open Form' THEN person_hash END) as form_openers,
            COUNT(DISTINCT CASE WHEN action_type = 'Submit' THEN person_hash END) as submitters,
            COUNT(DISTINCT CASE WHEN action_type = 'Open Invite' THEN person_hash END) as invite_openers,
            COUNT(DISTINCT CASE WHEN action_type = 'Send Invite' THEN person_hash END) as invite_senders
        FROM events
    """).fetchone()

    total_uv, readers, likers, form_openers, submitters, invite_openers, invite_senders = funnel

    # Funnel 1: Story Engagement
    r = 1
    r = _write_funnel(ws, r, "STORY ENGAGEMENT FUNNEL", [
        ("Total Visitors", total_uv),
        ("Read a Story", readers),
        ("Liked a Story", likers),
    ])
    r += 1

    # Funnel 2: Story Creation
    r = _write_funnel(ws, r, "STORY CREATION FUNNEL", [
        ("Total Visitors", total_uv),
        ("Opened Submission Form", form_openers),
        ("Submitted a Story", submitters),
    ])
    r += 1

    # Funnel 3: Invite Conversion
    r = _write_funnel(ws, r, "INVITE CONVERSION FUNNEL", [
        ("Total Visitors", total_uv),
        ("Opened Invite Dialog", invite_openers),
        ("Sent an Invite", invite_senders),
    ])

    finalize_sheet(ws, freeze_row=1)
    log("  Tab 7: Conversion Funnels")


def build_read_behaviour(wb, con):
    """Tab 8: Read Behaviour — aggregated read duration and follow-up analysis."""
    ws = wb.create_sheet("Read Behaviour")
    ws.sheet_properties.tabColor = GRAY_IV

    # Check if read_duration_sec column exists
    evt_cols = [rc[0] for rc in con.execute("DESCRIBE events").fetchall()]
    if 'read_duration_sec' not in evt_cols:
        ws.cell(row=1, column=1, value="No read behaviour data available (read_duration_sec missing)")
        log("  Tab 8: Read Behaviour (skipped — no data)")
        return

    # KPI query
    kpi = con.execute("""
        SELECT
            COUNT(*) as total_reads,
            COUNT(read_duration_sec) as reads_with_duration,
            COUNT(*) - COUNT(read_duration_sec) as reads_without_duration,
            ROUND(MEDIAN(read_duration_sec), 1) as median_duration,
            ROUND(AVG(read_duration_sec), 1) as avg_duration
        FROM events
        WHERE action_type = 'Read'
    """).fetchone()

    if not kpi or kpi[0] == 0:
        ws.cell(row=1, column=1, value="No read events found")
        log("  Tab 8: Read Behaviour (skipped — no reads)")
        return

    total_reads, with_dur, without_dur, median_dur, avg_dur = kpi

    # --- Section 1: KPI Overview ---
    r = 1
    write_section_header(ws, r, "READ BEHAVIOUR OVERVIEW", 5); r += 1
    write_kpi_row(ws, r, "Total Reads", total_reads, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Reads with Duration", with_dur, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Reads without Duration (session end)", without_dur, fmt=NUM_FMT_INT); r += 1
    write_kpi_row(ws, r, "Median Read Duration (s)", median_dur, fmt=NUM_FMT_RATIO); r += 1
    write_kpi_row(ws, r, "Avg Read Duration (s)", avg_dur, fmt=NUM_FMT_RATIO); r += 1
    r += 1

    # --- Section 2: Duration by Follow-up Action ---
    write_section_header(ws, r, "DURATION BY FOLLOW-UP ACTION", 5); r += 1
    s2_headers = ["Follow-up Action", "Count", "% of Reads", "Median Duration (s)", "Avg Duration (s)"]
    write_header_row(ws, r, s2_headers); r += 1

    action_stats = con.execute("""
        SELECT
            read_next_action as action,
            COUNT(*) as cnt,
            ROUND(100.0 * COUNT(*) / SUM(COUNT(*)) OVER(), 1) as pct,
            ROUND(MEDIAN(read_duration_sec), 1) as median_dur,
            ROUND(AVG(read_duration_sec), 1) as avg_dur
        FROM events
        WHERE action_type = 'Read'
        GROUP BY 1
        ORDER BY cnt DESC
    """).fetchall()

    for ai, (action, cnt, pct, med, avg) in enumerate(action_stats):
        fill = ALT_FILL if ai % 2 == 1 else None
        cell = ws.cell(row=r, column=1, value=action)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=2, value=cnt)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_INT
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=3, value=pct / 100.0 if pct else 0)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_PCT
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=4, value=med)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_RATIO
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=5, value=avg)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_RATIO
        if fill:
            cell.fill = fill
        r += 1

    r += 1

    # --- Section 3: Duration Distribution ---
    write_section_header(ws, r, "DURATION DISTRIBUTION", 3); r += 1
    s3_headers = ["Bucket", "Count", "% of Reads"]
    write_header_row(ws, r, s3_headers); r += 1

    bucket_rows = con.execute("""
        SELECT
            bucket,
            cnt,
            ROUND(100.0 * cnt / SUM(cnt) OVER(), 1) as pct
        FROM (
            SELECT
                CASE
                    WHEN read_duration_sec IS NULL THEN 'Unknown (session end)'
                    WHEN read_duration_sec < 5 THEN '< 5s'
                    WHEN read_duration_sec < 15 THEN '5 - 15s'
                    WHEN read_duration_sec < 30 THEN '15 - 30s'
                    WHEN read_duration_sec < 60 THEN '30 - 60s'
                    WHEN read_duration_sec < 300 THEN '1 - 5 min'
                    ELSE '>= 5 min (capped)'
                END as bucket,
                CASE
                    WHEN read_duration_sec IS NULL THEN 7
                    WHEN read_duration_sec < 5 THEN 1
                    WHEN read_duration_sec < 15 THEN 2
                    WHEN read_duration_sec < 30 THEN 3
                    WHEN read_duration_sec < 60 THEN 4
                    WHEN read_duration_sec < 300 THEN 5
                    ELSE 6
                END as sort_order,
                COUNT(*) as cnt
            FROM events
            WHERE action_type = 'Read'
            GROUP BY 1, 2
        )
        ORDER BY sort_order
    """).fetchall()

    for bi, (bucket, cnt, pct) in enumerate(bucket_rows):
        fill = ALT_FILL if bi % 2 == 1 else None
        cell = ws.cell(row=r, column=1, value=bucket)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=2, value=cnt)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_INT
        if fill:
            cell.fill = fill
        cell = ws.cell(row=r, column=3, value=pct / 100.0 if pct else 0)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_PCT
        if fill:
            cell.fill = fill
        r += 1

    r += 1

    # --- Section 4: Duration × Follow-up Cross-Tab ---
    write_section_header(ws, r, "DURATION × FOLLOW-UP ACTION (% of bucket leading to each action)", 8); r += 1

    # Get the distinct actions for column headers
    actions = con.execute("""
        SELECT DISTINCT read_next_action as action
        FROM events
        WHERE action_type = 'Read'
        ORDER BY action
    """).fetchall()
    action_names = [a[0] for a in actions]

    cross_headers = ["Duration Bucket"] + action_names + ["Total"]
    write_header_row(ws, r, cross_headers); r += 1

    # Query cross-tab data
    cross_data = con.execute("""
        SELECT
            bucket,
            sort_order,
            action,
            cnt,
            total_in_bucket
        FROM (
            SELECT
                CASE
                    WHEN read_duration_sec IS NULL THEN 'Unknown (session end)'
                    WHEN read_duration_sec < 5 THEN '< 5s'
                    WHEN read_duration_sec < 15 THEN '5 - 15s'
                    WHEN read_duration_sec < 30 THEN '15 - 30s'
                    WHEN read_duration_sec < 60 THEN '30 - 60s'
                    WHEN read_duration_sec < 300 THEN '1 - 5 min'
                    ELSE '>= 5 min (capped)'
                END as bucket,
                CASE
                    WHEN read_duration_sec IS NULL THEN 7
                    WHEN read_duration_sec < 5 THEN 1
                    WHEN read_duration_sec < 15 THEN 2
                    WHEN read_duration_sec < 30 THEN 3
                    WHEN read_duration_sec < 60 THEN 4
                    WHEN read_duration_sec < 300 THEN 5
                    ELSE 6
                END as sort_order,
                read_next_action as action,
                COUNT(*) as cnt,
                SUM(COUNT(*)) OVER (PARTITION BY
                    CASE
                        WHEN read_duration_sec IS NULL THEN 7
                        WHEN read_duration_sec < 5 THEN 1
                        WHEN read_duration_sec < 15 THEN 2
                        WHEN read_duration_sec < 30 THEN 3
                        WHEN read_duration_sec < 60 THEN 4
                        WHEN read_duration_sec < 300 THEN 5
                        ELSE 6
                    END
                ) as total_in_bucket
            FROM events
            WHERE action_type = 'Read'
            GROUP BY 1, 2, 3
        )
        ORDER BY sort_order, action
    """).fetchall()

    # Build lookup: (bucket, action) -> (cnt, total)
    cross_lookup = {}
    bucket_totals = {}
    for bucket, sort_order, action, cnt, total in cross_data:
        cross_lookup[(bucket, action)] = cnt
        bucket_totals[bucket] = total

    # Ordered buckets
    ordered_buckets = []
    seen = set()
    for bucket, sort_order, *_ in cross_data:
        if bucket not in seen:
            ordered_buckets.append(bucket)
            seen.add(bucket)

    for bi, bucket in enumerate(ordered_buckets):
        fill = ALT_FILL if bi % 2 == 1 else None
        total = bucket_totals.get(bucket, 0)
        # A: bucket label
        cell = ws.cell(row=r, column=1, value=bucket)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        # Action columns: % of bucket
        for ai, action in enumerate(action_names):
            cnt = cross_lookup.get((bucket, action), 0)
            pct = cnt / total if total > 0 else 0
            cell = ws.cell(row=r, column=ai + 2, value=pct)
            cell.border = THIN_BORDER
            cell.number_format = NUM_FMT_PCT
            if fill:
                cell.fill = fill
        # Total column
        cell = ws.cell(row=r, column=len(action_names) + 2, value=total)
        cell.border = THIN_BORDER
        cell.number_format = NUM_FMT_INT
        if fill:
            cell.fill = fill
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    for ci in range(6, len(action_names) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    finalize_sheet(ws)
    log("  Tab 8: Read Behaviour")


def build_glossary(wb):
    """Tab 9: Glossary — metric definitions and context for report readers."""
    ws = wb.create_sheet("Glossary")
    ws.sheet_properties.tabColor = GRAY_IV
    ws.sheet_view.showGridLines = False

    TITLE_FONT = Font(bold=True, color=GRAY_VI, size=13)
    HEADING_FONT = Font(bold=True, color=GRAY_VI, size=11)
    BODY_FONT = Font(color=GRAY_VI, size=10)
    TERM_FONT = Font(bold=True, color=BLACK, size=10)

    r = 1
    ws.cell(row=r, column=1, value="Glossary").font = TITLE_FONT
    r += 2

    def heading(text):
        nonlocal r
        ws.cell(row=r, column=1, value=text).font = HEADING_FONT
        r += 1

    def term(name, definition):
        nonlocal r
        ws.cell(row=r, column=1, value=name).font = TERM_FONT
        ws.cell(row=r, column=1).alignment = Alignment(indent=1)
        ws.cell(row=r, column=2, value=definition).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        r += 1

    def blank():
        nonlocal r
        r += 1

    # --- Metrics ---
    heading("Metrics")
    term("Total Clicks", "All tracked user interactions on the platform (Reads, Likes, Form actions, Invites).")
    term("Unique Visitors", "Number of distinct users who performed at least one click.")
    term("Unique Sessions", "Number of distinct browsing sessions (one user can have multiple sessions).")
    term("Engaged Visitors", "Users who performed at least one Read or Like.")
    term("Engagements", "Total count of Read and Like actions combined.")
    term("Like Rate", "Likes divided by Reads (or Unique Readers). Shows how often readers appreciate a story.")
    term("Reads/Day", "Total Reads divided by story lifespan in days. Normalizes for story age.")
    blank()

    # --- Action Types ---
    heading("Action Types")
    term("Read", "User opened or expanded a story to view its content.")
    term("Like", "User liked a story.")
    term("Open Form", "User opened the story submission form.")
    term("Submit", "User submitted a new story.")
    term("Cancel", "User closed the submission form without submitting.")
    term("Delete", "User confirmed deletion of a story.")
    term("Open Invite", "User opened the colleague invite dialog.")
    term("Send Invite", "User sent an invitation to colleagues.")
    blank()

    # --- Click Categories ---
    heading("Click Categories (Executive Summary)")
    term("Engagement", "Read + Like. Represents genuine story consumption and appreciation.")
    term("Invite", "Open Invite + Send Invite. Represents viral/sharing behavior.")
    term("Submission", "Open Form + Submit + Delete. Represents content creation activity.")
    blank()

    # --- Engagement Definition ---
    heading("Engagement Definition")
    ws.cell(row=r, column=1).font = BODY_FONT
    ws.cell(row=r, column=1, value=(
        "Throughout this report, 'engagement' is defined as Read + Like actions. "
        "These represent meaningful interaction with story content. Other click types "
        "(form actions, invites) are tracked separately as they represent different user journeys."
    )).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    # --- Funnels ---
    heading("Conversion Funnels")
    term("Story Engagement", "Total Visitors > Read a Story > Liked a Story")
    term("Story Creation", "Total Visitors > Opened Submission Form > Submitted a Story")
    term("Invite Conversion", "Total Visitors > Opened Invite Dialog > Sent an Invite")
    term("% of Previous Step", "Percentage of users who progressed from the immediately preceding step.")
    term("% of First Step", "Percentage relative to Total Visitors (the funnel entry point).")
    blank()

    # --- Data Notes ---
    heading("Data Notes")
    term("Timezone", "All timestamps and session dates are in Central European Time (CET/CEST).")
    term("Deleted Stories", "Stories removed from the platform. Historical engagement data up to the deletion date is preserved.")
    term("3Keys", "Up to three category tags assigned to each story. Used to analyze which topics resonate most.")

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    log("  Tab 8: Glossary")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Generate CampaignWe XLSX Report")
    parser.add_argument("--output", type=str, default=None,
                        help="Output XLSX path (default: output/campaignwe_report.xlsx)")
    parser.add_argument("--date-from", type=str, default=None,
                        help="Start date filter (YYYY-MM-DD)")
    parser.add_argument("--date-to", type=str, default=None,
                        help="End date filter (YYYY-MM-DD)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    output_dir = script_dir / "output"
    events_path = output_dir / "events_anonymized.parquet"
    metadata_path = output_dir / "story_metadata.parquet"

    today = date.today().strftime("%Y_%m_%d")
    output_path = Path(args.output) if args.output else output_dir / f"campaignwe_report_{today}.xlsx"

    log("=" * 64)
    log("  CampaignWe XLSX Report Generator")
    log("=" * 64)
    log()
    log("Loading data...")

    con, cols = load_data(events_path, metadata_path, args.date_from, args.date_to)

    log()
    log("Building report tabs...")
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    build_executive_summary(wb, con, cols)
    build_weekly_trend(wb, con)
    build_story_performance(wb, con)
    build_key_performance(wb, con)
    build_division_engagement(wb, con, cols)
    build_region_engagement(wb, con, cols)
    build_hourly_weekday(wb, con)
    build_conversion_funnel(wb, con)
    build_read_behaviour(wb, con)
    build_glossary(wb)

    log()
    log(f"Saving to {output_path}...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    size_kb = output_path.stat().st_size / 1024
    log(f"  Done! ({size_kb:.1f} KB)")
    log()

    con.close()


if __name__ == "__main__":
    main()
